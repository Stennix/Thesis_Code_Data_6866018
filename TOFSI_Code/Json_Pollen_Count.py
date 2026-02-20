"""
Script to extract pollen counts in JSON files.

Outputs:
- Global totals
- Per slide totals
- Excel matrix
- Detailed pollen detection locator

Script generated using AI prompts
"""

import json
import os
from collections import Counter, defaultdict
from glob import glob
from openpyxl import Workbook

############################################################
# CONFIG
############################################################

json_root_folder = r"YOUR_JSON_ROOT_FOLDER"
output_excel_path = r"EXCEL_OUTPUT"

# --- LOCATOR SETTINGS ---
ENABLE_LOCATOR = True
TARGET_LABELS = ["YOUR", "TARGETED", "LABELS"]   # Example: ["Jug", "Ole"]

############################################################
# FIND JSON FILES
############################################################

json_files = glob(os.path.join(json_root_folder, "**", "*.json"), recursive=True)
print(f"Found {len(json_files)} JSON files")

############################################################
# COUNTERS
############################################################

# Global counters
global_counter = Counter()
global_total = 0

# Per slide counters
folder_counters = defaultdict(Counter)
folder_totals = defaultdict(int)

# All unique labels (for Excel)
all_labels = set()

# Locator structure:
# locator[label][slide][json_file] = count
locator = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))

############################################################
# PROCESS FILES
############################################################

for jf in json_files:

    try:
        with open(jf, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        print(f"⚠️ Could not read {jf}: {e}")
        continue

    if "shapes" not in data:
        continue

    # Slide name = first folder under root
    relative_path = os.path.relpath(jf, json_root_folder)
    slide_name = relative_path.split(os.sep)[0]

    for shape in data["shapes"]:
        label = shape.get("label", "UNKNOWN")

        # --- GLOBAL ---
        global_counter[label] += 1
        global_total += 1

        # --- PER SLIDE ---
        folder_counters[slide_name][label] += 1
        folder_totals[slide_name] += 1

        # Track all labels for Excel
        all_labels.add(label)

        # --- LOCATOR ---
        if ENABLE_LOCATOR and label in TARGET_LABELS:
            json_name = os.path.basename(jf)
            locator[label][slide_name][json_name] += 1

############################################################
# GLOBAL SUMMARY
############################################################

print("\n==============================")
print("GLOBAL TOTALS")
print("==============================")

for label, count in global_counter.most_common():
    print(f"{label:20s} {count}")

print(f"\nTotal objects (incl. Nonpollen): {global_total}")
nonpollen = global_counter.get("Nonpollen", 0)
print(f"Pollen objects only: {global_total - nonpollen}")

############################################################
# PER SLIDE SUMMARY
############################################################

print("\n==============================")
print("PER SLIDE TOTALS")
print("==============================")

for slide in sorted(folder_counters.keys()):
    print(f"\nSlide: {slide}")
    print(f"Total objects: {folder_totals[slide]}")
    nonpollen = folder_counters[slide].get("Nonpollen", 0)
    print(f"Pollen only: {folder_totals[slide] - nonpollen}")

    for label, count in folder_counters[slide].most_common():
        print(f"  {label:18s} {count}")

############################################################
# LOCATOR OUTPUT (Per pollen → per slide → per JSON)
############################################################

if ENABLE_LOCATOR and TARGET_LABELS:

    print("\n==============================")
    print("LOCATOR RESULTS (Detailed)")
    print("==============================")

    for label in TARGET_LABELS:

        print(f"\n\n===== LABEL: {label} =====")

        if label not in locator:
            print("No detections found.")
            continue

        for slide in sorted(locator[label].keys()):
            print(f"\nSlide: {slide}")

            for json_file, count in sorted(locator[label][slide].items()):
                print(f"  {json_file:30s} {count}")

############################################################
# CREATE EXCEL MATRIX
############################################################

print("\nCreating Excel file...")

wb = Workbook()
ws = wb.active
ws.title = "Pollen Counts"

sorted_slides = sorted(folder_counters.keys())
sorted_labels = sorted(all_labels)

# Header row
ws.cell(row=1, column=1, value="Label")

for col_index, slide in enumerate(sorted_slides, start=2):
    ws.cell(row=1, column=col_index, value=slide)

# Fill matrix
for row_index, label in enumerate(sorted_labels, start=2):
    ws.cell(row=row_index, column=1, value=label)

    for col_index, slide in enumerate(sorted_slides, start=2):
        count = folder_counters[slide].get(label, 0)
        ws.cell(row=row_index, column=col_index, value=count)

wb.save(output_excel_path)

print(f"Excel file saved to:\n{output_excel_path}")